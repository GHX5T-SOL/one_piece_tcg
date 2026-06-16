#!/usr/bin/env python3
"""Refresh the OP CPT / The Vault Room business workspace from the current Collectr export.

This script intentionally writes durable markdown, CSV, and XLSX artifacts so
the repo can act as the OP CPT operating brain. The public-facing brand is
The Vault Room. This script does not call external APIs; live price research
still happens per card through the pricing protocol.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit("openpyxl is required; use the bundled Codex Python runtime") from exc


ROOT = Path(__file__).resolve().parents[1]
TODAY = date.today().isoformat()
EXPORT_SOURCE = Path("/Users/mx/Downloads/export.csv")

COLLECTR_PROFILE = "https://app.getcollectr.com/showcase/profile/ec11ff42-0dfe-4826-9386-d198a3174e67"
GITHUB_REPO = "https://github.com/GHX5T-SOL/one_piece_tcg"
LIVE_SITE = "https://op-cpt.vercel.app"
SHARED_CHAT = "https://chatgpt.com/share/6a294dd4-ea44-8326-85c8-127becfd71fa"
CONSIGNMENT_DRIVE = "https://drive.google.com/drive/u/0/mobile/folders/1CUekY1s7sZ5lSScpQD7-5kFkANZNAgRR?usp=drive_link"
PRICE_BIBLE_REPORT = Path("/Users/mx/Desktop/deep-research-report.md")
PUBLIC_BRAND = "The Vault Room"
INTERNAL_WORKSPACE = "OP CPT"
INSTAGRAM_HANDLE = "@thevaultroom.cpt"
INSTAGRAM_URL = "https://instagram.com/thevaultroom.cpt"
WHATSAPP_COMMUNITY = "https://chat.whatsapp.com/LgAy8Q0NgVp9NcU7oJXBDy"
BRAND_TAGLINE = "Cards. Collectibles. Grails."
BRAND_COMMUNITY_LINE = "Cape Town Collector Community"


MASTER_COLUMNS = [
    "item_id",
    "source_row_id",
    "ownership_status",
    "inventory_status",
    "card_name",
    "card_number",
    "set_code",
    "set_name",
    "product_name",
    "rarity",
    "variant",
    "finish",
    "language",
    "condition",
    "grade_company",
    "grade",
    "cert_number",
    "quantity",
    "sealed_or_single",
    "raw_or_slab",
    "cost_basis_zar",
    "cost_status",
    "collectr_market_zar",
    "research_market_usd",
    "research_market_zar",
    "quick_sale_zar",
    "fair_market_zar",
    "patient_local_ask_zar",
    "event_ask_zar",
    "online_ask_usd",
    "online_ask_zar",
    "stretch_ask_zar",
    "walkaway_minimum_zar",
    "expected_net_local_zar",
    "expected_net_ebay_zar",
    "expected_net_cardtrader_zar",
    "expected_net_cardmarket_zar",
    "expected_net_tcgplayer_zar",
    "grading_candidate_score",
    "psa10_estimate_usd",
    "bgs10_estimate_usd",
    "cgc10_estimate_usd",
    "availability_score",
    "liquidity_score",
    "trend_score",
    "confidence_rating",
    "last_researched_at",
    "source_links",
    "notes",
]


def ensure_dirs() -> None:
    for folder in [
        "market_research",
        "pricing",
        "inventory/raw",
        "website",
        "content",
        "memory",
        "scripts",
    ]:
        (ROOT / folder).mkdir(parents=True, exist_ok=True)


def clean_money(value: str | None) -> float:
    if not value:
        return 0.0
    value = value.replace(",", "").strip()
    try:
        return float(value)
    except ValueError:
        return 0.0


def clean_int(value: str | None) -> int:
    if not value:
        return 0
    try:
        return int(float(value.replace(",", "")))
    except ValueError:
        return 0


def slugify(text: str, fallback: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return slug[:80] or fallback


def set_code(card_number: str) -> str:
    match = re.match(r"([A-Z]{1,4}\d{0,2})[- ]", card_number or "")
    return match.group(1) if match else ""


def infer_language(product_name: str, set_name: str) -> str:
    combined = f"{product_name} {set_name}".lower()
    if "(jp)" in combined or "japanese" in combined:
        return "Japanese"
    if "(en)" in combined or "english" in combined:
        return "English"
    return "Unknown"


def infer_variant(product_name: str, rarity: str, variance: str) -> str:
    text = f"{product_name} {rarity} {variance}".lower()
    tags: list[str] = []
    if "alternate art" in text or "alt art" in text:
        tags.append("alternate_art")
    if "parallel" in text:
        tags.append("parallel")
    if "sp)" in text or "(sp" in text or "special" in text:
        tags.append("sp")
    if "manga" in text:
        tags.append("manga")
    if "winner" in text:
        tags.append("winner")
    if "event" in text:
        tags.append("event")
    if "promo" in text or rarity.upper() in {"P", "PR", "PROMO"}:
        tags.append("promo")
    if "pre-release" in text or "prerelease" in text:
        tags.append("pre_release")
    return ";".join(dict.fromkeys(tags)) or "base_or_standard"


def infer_sealed(product_name: str, card_number: str, rarity: str) -> str:
    text = f"{product_name} {card_number} {rarity}".lower()
    sealed_terms = [
        "sealed",
        "booster",
        "sleeved booster",
        "deck set",
        "premium card collection",
        "starter deck",
    ]
    if any(term in text for term in sealed_terms) and not re.search(r"\b(op|st|eb|p)-?\d", card_number.lower()):
        return "sealed_product"
    if "premium card collection" in text and not card_number:
        return "sealed_product"
    return "single"


def parse_grade(grade: str) -> tuple[str, str, str]:
    grade = (grade or "").strip()
    if not grade or grade.lower() == "ungraded":
        return "RAW", "", "raw"
    company = grade.split()[0].replace(".", "").upper()
    return company, grade, "slab"


def read_export() -> list[dict[str, str]]:
    if not EXPORT_SOURCE.exists():
        return []
    target = ROOT / "inventory/raw/export.csv"
    target.write_bytes(EXPORT_SOURCE.read_bytes())
    with EXPORT_SOURCE.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def build_master(rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    master: list[dict[str, Any]] = []
    category_counter: Counter[str] = Counter()
    grade_counter: Counter[str] = Counter()
    rarity_counter: Counter[str] = Counter()
    set_counter: Counter[str] = Counter()
    value_by_set: defaultdict[str, float] = defaultdict(float)
    total_qty = 0
    market_total = 0.0
    cost_total = 0.0
    zero_cost_rows = 0
    recorded_cost_rows = 0
    slabbed_rows = 0
    raw_rows = 0
    sealed_rows = 0

    for index, row in enumerate(rows, start=1):
        product = row.get("Product Name", "").strip()
        card_no = row.get("Card Number", "").strip()
        set_name = row.get("Set", "").strip()
        rarity = row.get("Rarity", "").strip()
        variance = row.get("Variance", "").strip()
        grade_company, grade_text, raw_or_slab = parse_grade(row.get("Grade", ""))
        quantity = clean_int(row.get("Quantity")) or 1
        market = clean_money(row.get("Market Price (As of 2026-06-08)"))
        cost = clean_money(row.get("Average Cost Paid"))
        sealed = infer_sealed(product, card_no, rarity)
        item_id = f"{index:04d}-{slugify(product + '-' + card_no, f'row-{index}')}"
        row_market_total = market * quantity
        row_cost_total = cost * quantity

        master_row = {
            "item_id": item_id,
            "source_row_id": index,
            "ownership_status": "in_hand",
            "inventory_status": "available_for_pricing",
            "card_name": product,
            "card_number": card_no,
            "set_code": set_code(card_no),
            "set_name": set_name,
            "product_name": product,
            "rarity": rarity,
            "variant": infer_variant(product, rarity, variance),
            "finish": variance,
            "language": infer_language(product, set_name),
            "condition": row.get("Card Condition", "").strip() or "Unknown",
            "grade_company": grade_company,
            "grade": grade_text,
            "cert_number": "",
            "quantity": quantity,
            "sealed_or_single": sealed,
            "raw_or_slab": raw_or_slab,
            "cost_basis_zar": f"{row_cost_total:.2f}" if cost else "",
            "cost_status": "recorded" if cost else "unknown_or_box_pull",
            "collectr_market_zar": f"{row_market_total:.2f}" if market else "",
            "research_market_usd": "",
            "research_market_zar": "",
            "quick_sale_zar": "",
            "fair_market_zar": "",
            "patient_local_ask_zar": "",
            "event_ask_zar": "",
            "online_ask_usd": "",
            "online_ask_zar": "",
            "stretch_ask_zar": "",
            "walkaway_minimum_zar": "",
            "expected_net_local_zar": "",
            "expected_net_ebay_zar": "",
            "expected_net_cardtrader_zar": "",
            "expected_net_cardmarket_zar": "",
            "expected_net_tcgplayer_zar": "",
            "grading_candidate_score": "",
            "psa10_estimate_usd": "",
            "bgs10_estimate_usd": "",
            "cgc10_estimate_usd": "",
            "availability_score": "",
            "liquidity_score": "",
            "trend_score": "",
            "confidence_rating": "pending_research",
            "last_researched_at": "",
            "source_links": COLLECTR_PROFILE,
            "notes": "Imported from Collectr export. Collectr value is baseline only; research before quoting.",
        }
        master.append(master_row)

        category_counter[row.get("Category", "").strip()] += 1
        grade_counter[row.get("Grade", "").strip() or "Unknown"] += 1
        rarity_counter[rarity or "Unknown"] += 1
        set_counter[set_name or "Unknown"] += 1
        value_by_set[set_name or "Unknown"] += row_market_total
        total_qty += quantity
        market_total += row_market_total
        cost_total += row_cost_total
        zero_cost_rows += int(cost == 0)
        recorded_cost_rows += int(cost > 0)
        slabbed_rows += int(raw_or_slab == "slab")
        raw_rows += int(raw_or_slab == "raw")
        sealed_rows += int(sealed == "sealed_product")

    summary = {
        "generated_at": TODAY,
        "source_file": str(EXPORT_SOURCE),
        "raw_export_copied_to": "inventory/raw/export.csv" if rows else "",
        "rows": len(rows),
        "total_quantity": total_qty,
        "category_rows": dict(category_counter),
        "grade_rows": dict(grade_counter),
        "rarity_rows": dict(rarity_counter),
        "top_sets_by_rows": dict(set_counter.most_common(12)),
        "top_sets_by_collectr_market_zar": dict(sorted(value_by_set.items(), key=lambda item: item[1], reverse=True)[:12]),
        "zero_cost_rows": zero_cost_rows,
        "recorded_cost_rows": recorded_cost_rows,
        "csv_market_value_baseline_zar": round(market_total, 2),
        "recorded_cost_basis_zar": round(cost_total, 2),
        "watchlist_rows": sum(1 for r in rows if (r.get("Watchlist") or "").lower() == "true"),
        "raw_rows": raw_rows,
        "slabbed_rows": slabbed_rows,
        "sealed_product_rows": sealed_rows,
        "collectr_warning": "Collectr values are portfolio baselines only and are not sell prices.",
    }
    return master, summary


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def write_xlsx(master: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Master Inventory"
    ws.append(MASTER_COLUMNS)
    header_fill = PatternFill(fill_type="solid", fgColor="0E7490")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in master:
        ws.append([row.get(column, "") for column in MASTER_COLUMNS])
    ws.freeze_panes = "A2"
    for idx, column in enumerate(MASTER_COLUMNS, start=1):
        width = min(max(len(column) + 2, 12), 32)
        ws.column_dimensions[get_column_letter(idx)].width = width

    summary_ws = workbook.create_sheet("Baseline Summary")
    summary_ws.append(["metric", "value"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    for key, value in summary.items():
        summary_ws.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
    summary_ws.column_dimensions["A"].width = 36
    summary_ws.column_dimensions["B"].width = 90
    workbook.save(ROOT / "inventory/master_inventory.xlsx")


def write_logs() -> None:
    logs = {
        "inventory/pricing_research_log.csv": [
            "researched_at",
            "item_id",
            "card_name",
            "card_number",
            "platform",
            "source_url",
            "currency",
            "price",
            "condition",
            "language",
            "variant",
            "signal_type",
            "included_in_comp_stack",
            "exclusion_reason",
            "notes",
        ],
        "inventory/sales_log.csv": [
            "sold_at",
            "item_id",
            "card_name",
            "channel",
            "gross_zar",
            "fees_zar",
            "shipping_zar",
            "net_zar",
            "buyer_handle",
            "payment_status",
            "delivery_status",
            "notes",
        ],
        "inventory/purchase_log.csv": [
            "purchased_at",
            "item_id",
            "card_name",
            "source",
            "gross_zar",
            "fees_zar",
            "shipping_zar",
            "landed_cost_zar",
            "ownership_status",
            "notes",
        ],
        "inventory/consignment_intake.csv": [
            "intake_at",
            "consignment_id",
            "owner_alias",
            "card_name",
            "card_number",
            "condition",
            "researched_market_zar",
            "consignment_baseline_percent",
            "owner_expected_zar",
            "op_cpt_target_margin_zar",
            "status",
            "notes",
        ],
        "inventory/watchlist.csv": [
            "created_at",
            "card_name",
            "card_number",
            "target_variant",
            "target_grade",
            "ideal_bid_zar",
            "max_bid_zar",
            "walkaway_zar",
            "reason",
            "source_url",
            "status",
        ],
    }
    for relative, columns in logs.items():
        write_csv(ROOT / relative, [], columns)


def write_markdown(path: str, body: str) -> None:
    target = ROOT / path
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(body.strip() + "\n", encoding="utf-8")


def source_ledger() -> str:
    rows = [
        ("Official One Piece Card Game", "https://en.onepiece-cardgame.com/", "official", "Identity, products, rules, events", "No market prices", 5, "Primary official English source."),
        ("Official products", "https://en.onepiece-cardgame.com/products/", "official", "Release cadence and sealed-product identity", "Dynamic pages require recheck", 5, "OP-16 product page verified on 2026-06-16."),
        ("OP-16 The Time of Battle", "https://en.onepiece-cardgame.com/products/op16.html", "official/product", "OP-16 release date, MSRP, rarity mix", "Does not price singles", 5, "Official page states release date June 12, 2026 and 126+1 card types."),
        ("Official card list", "https://en.onepiece-cardgame.com/cardlist/", "official/card-data", "Exact card identity, text, rarity, set", "Must filter exact set/language", 5, "Use before pricing card numbers."),
        ("Official restrictions", "https://en.onepiece-cardgame.com/rules/restriction/", "official/rules", "Ban, banned-pair, restricted status", "Meta impact needs tournament data", 5, "Effective April 10, 2026 restriction notice verified."),
        ("Card Ladder OnePiece", "https://www.cardladder.com/ladder?category=OnePiece", "market/graded", "High-end graded trend, last sold, ladder movement", "PRO limits and exact-card matching required", 4, "Required by v2 pricing protocol for serious cards."),
        ("TCGplayer", "https://www.tcgplayer.com/", "market/us-retail", "US retail, market price, listing depth", "Sales history may be gated; verify variant", 4, ""),
        ("CardTrader", "https://www.cardtrader.com/", "market/global", "Global availability and CardTrader Zero/Ready context", "Fees and shipping change landed cost", 4, ""),
        ("Cardmarket One Piece", "https://www.cardmarket.com/en/OnePiece", "market/eu", "EU floor, trend, active supply", "EU prices do not equal Cape Town prices", 4, ""),
        ("PriceCharting", "https://www.pricecharting.com/", "market/history", "Raw/graded historical sale snapshots", "SKU matching can be wrong", 3, ""),
        ("eBay", "https://www.ebay.com/", "market/global", "Active/sold listings and rare-card scarcity", "Scams, title stuffing, accepted-offer opacity", 4, "Use solds and active listings with filtering."),
        ("PSA", "https://www.psacard.com/", "grading", "Cert verification, population, APR", "Access limits and pop changes", 4, ""),
        ("BGS", "https://www.beckett.com/grading", "grading", "BGS grading/pop context", "Exact pop checks may be gated", 4, ""),
        ("CGC Cards", "https://www.cgccards.com/", "grading", "CGC cert and pop context", "CGC 10 and Pristine must be separated", 4, ""),
        ("GemRate", "https://www.gemrate.com/", "grading/market", "Universal pop and grading-volume trends", "Not every card has complete detail", 4, "Universal search and daily grading trends verified."),
        ("Alt", "https://www.alt.xyz/", "market/portfolio", "Slab portfolio and premium collectible sentiment", "Coverage may be incomplete", 3, ""),
        ("Market Movers", "https://www.marketmoversapp.com/", "market/data", "Broader card sales charts and alerts", "Subscription/access limits", 3, ""),
        ("130point", "https://130point.com/", "market/sold-helper", "Accepted-offer/eBay helper", "Availability varies", 3, ""),
        ("Limitless One Piece", "https://onepiece.limitlesstcg.com/", "competitive/card-data", "Variants, tournament data, TCGplayer/Cardmarket surfaces", "Competitive demand not equal collector demand", 4, ""),
        ("OnePiece.gg", "https://onepiece.gg/", "competitive/content", "Meta summaries and deck guides", "Editorial and may lag official updates", 3, ""),
        ("Nakama Decks", "https://nakamadecks.com/", "competitive/community", "Deck usage and leader/card demand", "Community data needs validation", 3, ""),
        ("Egman Events", "https://egmanevents.com/", "competitive/events", "Tournament results and event signal", "Coverage varies", 3, ""),
        ("Bandai Namco IR card business feature", "https://www.bandainamco.co.jp/en/ir/library/feature02_01_2025.html", "company/market", "Bandai card-business growth context", "Corporate framing, not secondary market", 5, "Bandai says growth was driven by One Piece Card Game."),
        ("PSA One Piece collecting basics", "https://www.psacard.com/info/tcg/one-piece-collecting-basics-guide", "grading/education", "PSA-recognized One Piece collecting categories", "Education, not pricing", 4, ""),
        ("PSA infrastructure investment", "https://www.psacard.com/articles/articleview/15715/200-million-investment-grading-experience", "grading/operations", "Turnaround and demand context", "Not One Piece-specific", 4, "May 14, 2026 service-level update verified."),
        ("Collectr", "https://app.getcollectr.com/", "portfolio/noisy", "Portfolio baseline and SKU discovery", "Never final sell price", 2, "Weak baseline only."),
        ("OP CPT live site", LIVE_SITE, "owned/site", "Current public site behavior", "Test before changing", 5, "Verified public shell on 2026-06-16."),
        ("OP CPT GitHub repo", GITHUB_REPO, "owned/repo", "Source of truth and backup", "Currently public", 5, "Public repo verified by gh and GitHub page."),
        ("The Vault Room Instagram", INSTAGRAM_URL, "owned/social", "Public social identity and content channel", "Handle may need in-app verification before launch claims", 5, INSTAGRAM_HANDLE),
        ("The Vault Room WhatsApp community", WHATSAPP_COMMUNITY, "owned/community", "Community join route for Cape Town collectors", "Invite links can expire or be replaced", 4, "Use as current public join link until changed."),
        ("Collectr portfolio", COLLECTR_PROFILE, "owned/portfolio", "Current portfolio source", "Collectr values need validation", 3, "Public profile title loaded on 2026-06-16."),
        ("Consignment Drive folder", CONSIGNMENT_DRIVE, "private/intake", "Consignment image intake", "Access and owner data must be handled carefully", 3, "Do not expose owner/private details in public listings."),
        ("Shared ChatGPT context", SHARED_CHAT, "context", "Prior OP CPT setup context", "Web fetch showed login shell only", 2, "Ask user to export/paste if full transcript is needed."),
        ("Vault Room poster example", "media/branding/vault-room-poster-example.jpg", "local/brand-asset", "Poster direction, community promise, visual tone", "Generated/reference asset; review text before public print", 4, "Provided by user on 2026-06-16."),
        ("Vault Room crest", "media/branding/vault-room-crest.jpg", "local/brand-asset", "Profile icon and crest reference", "Contains fan-symbol influences; keep public usage reviewed", 4, "Provided by user on 2026-06-16."),
        ("Vault Room primary logo", "media/branding/vault-room-primary-logo.jpg", "local/brand-asset", "Primary lockup reference", "Generated/reference asset; vector cleanup may be needed", 4, "Provided by user on 2026-06-16."),
        ("Vault Room brand board", "media/branding/vault-room-brand-board.jpg", "local/brand-asset", "Palette, lockups, concept, and brand pillars", "Generated/reference asset; public-safe cleanup needed before merch", 4, "Provided by user on 2026-06-16."),
    ]
    lines = [
        "# Source Ledger",
        "",
        f"Updated: {TODAY}",
        "",
        "Tier rule: official identity/rules first, then clean sold comps, then graded/pop sources, then social/local sentiment. Collectr is a weak baseline only.",
        "",
        "| date_accessed | source_name | url | source_type | what_it_is_good_for | limitations | trust_score_1_to_5 | notes |",
        "|---|---|---|---|---|---|---:|---|",
    ]
    for name, url, kind, good, limits, score, notes in rows:
        lines.append(f"| {TODAY} | {name} | {url} | {kind} | {good} | {limits} | {score} | {notes} |")
    return "\n".join(lines)


def write_docs(summary: dict[str, Any]) -> None:
    summary_json = json.dumps(summary, indent=2, ensure_ascii=False)
    write_markdown(
        "README_OP_CPT.md",
        f"""
# OP CPT / The Vault Room Business Workspace

The Vault Room is the public brand for the Cape Town collector community and trading-card business. OP CPT remains the internal workspace name for buying, selling, pricing, inventory, grading, consignment, trade days, tournaments, content, and the current website.

Brand identity:

- Public brand: {PUBLIC_BRAND}
- Tagline: {BRAND_TAGLINE}
- Community line: {BRAND_COMMUNITY_LINE}
- Instagram: {INSTAGRAM_HANDLE}
- WhatsApp community: {WHATSAPP_COMMUNITY}

Primary links:

- GitHub repo: {GITHUB_REPO}
- Live website: {LIVE_SITE}
- Collectr portfolio: {COLLECTR_PROFILE}
- Consignment Drive intake: {CONSIGNMENT_DRIVE}
- Shared ChatGPT context: {SHARED_CHAT}

## Operating Model

This repo is the persistent business brain. Use the files in `pricing/`, `inventory/`, `market_research/`, `content/`, `website/`, and `memory/` before answering future Vault Room or OP CPT questions.

For card prices, always follow `pricing/pricing_protocol.md`. Do not price from Collectr alone.

## Current Inventory Baseline

- Rows: {summary.get('rows', 0)}
- Quantity: {summary.get('total_quantity', 0)}
- Collectr CSV market baseline: R{summary.get('csv_market_value_baseline_zar', 0):,.2f}
- Recorded cost basis: R{summary.get('recorded_cost_basis_zar', 0):,.2f}
- Slabbed rows: {summary.get('slabbed_rows', 0)}
- Raw rows: {summary.get('raw_rows', 0)}
- Sealed-product rows: {summary.get('sealed_product_rows', 0)}

The market baseline is not a sell price. Every sale, bid, consignment, or grading decision needs fresh research.
""",
    )
    write_markdown(
        "mission.md",
        """
# Mission

Build The Vault Room into Cape Town's trusted collector community for cards, collectibles, and grails: premium sourcing, transparent pricing, rare-card access, grading intelligence, consignment, trade days, tournaments, and a public catalogue that earns buyer trust.

The business should feel collector-first, research-backed, and locally useful. It should not race to be the cheapest seller.
""",
    )
    write_markdown(
        "goals.md",
        """
# Goals

## Commercial

- Turn the current in-hand inventory into a priced, source-backed sales catalogue.
- Build a repeatable card-pricing workflow for raw, slabbed, sealed, and consignment items.
- Track cost basis, expected net, sale channel, and ROI.
- Build a patient high-ask pricing culture with clear walk-away minimums.

## Community

- Grow The Vault Room as a Cape Town collector community for trades, battles, pack rips, group buys, and trade days.
- Support tournaments, rankings, event recaps, and local trust.

## Website

- Evolve https://op-cpt.vercel.app into an enquiry-only catalogue before considering checkout.
- Keep public/private data separation strict.
""",
    )
    write_markdown(
        "operating_principles.md",
        """
# Operating Principles

- Never price from Collectr alone.
- Start high, negotiate down, and protect rare-card upside.
- Identify exact card, language, variant, condition, stamp, and grade before quoting.
- Separate raw value from grading upside.
- Quote in ZAR and USD for serious pricing.
- Local Cape Town in-hand availability can justify a premium.
- Protect reputation: no condition, grade, rarity, or origin misrepresentation.
- Keep secrets, customer data, payment details, and private addresses out of public files.
- Save decisions and evidence to the workspace so the next session starts from repo truth.
""",
    )
    write_markdown(
        "questions.md",
        f"""
# Questions

Non-blocking questions saved during bootstrap:

1. Should the GitHub repo be made private before future high-resolution scans, cert images, private cost-basis notes, and consignment-owner details are committed?
2. The shared ChatGPT context at {SHARED_CHAT} loaded only a login shell through web fetch. Please export or paste the transcript if it contains details not already captured here.
3. Confirm whether the public catalogue should show exact available quantities or only `available / reserved / sold` status.
4. Confirm whether consignment owner names should be stored as aliases only, even locally.
5. Confirm target margin for owned stock by tier: fast flip, fair sale, premium hold, and event/show sticker.
""",
    )
    write_markdown(
        "decision_log.md",
        f"""
# Decision Log

| date | decision | reason | impact |
|---|---|---|---|
| {TODAY} | Treat The Vault Room as the public brand and OP CPT as the internal workspace. | User supplied new branding, Instagram handle, WhatsApp community, logo, and poster examples. | Future docs, pricing, content, and website revamp work should use The Vault Room publicly while preserving OP CPT continuity. |
| {TODAY} | Treat OP CPT / The Vault Room as one operating workspace. | User supplied pricing protocol and business bootstrap prompt. | All future card pricing, inventory, content, website, and consignment work should route through this workspace. |
| {TODAY} | Use Collectr export as in-hand inventory baseline, not pricing truth. | Protocol says Collectr is weak only; export has market baseline but not sell prices. | Master inventory uses pending researched fields until live comps are checked. |
| {TODAY} | Make Card Ladder a required source for serious One Piece pricing. | Pricing Protocol v2 explicitly added Card Ladder. | High-end slabs, manga rares, prize cards, promos, and low-pop cards need Card Ladder checks when accessible. |
| {TODAY} | Keep current website enquiry-first. | Bootstrap prompt says checkout/payment can wait. | Website roadmap focuses on public catalogue, contact, scans, and admin workflow. |
| {TODAY} | Commit workspace artifacts despite public repo per user instruction, but never commit secrets. | User explicitly approved pushing everything and will make the repo private later. | Inventory export and generated sheets are included; secrets remain excluded. |
""",
    )
    write_markdown("source_ledger.md", source_ledger())
    write_markdown(
        "brand_identity.md",
        f"""
# Brand Identity

Updated: {TODAY}

## Current Public Brand

- Name: {PUBLIC_BRAND}
- Tagline: {BRAND_TAGLINE}
- Community line: {BRAND_COMMUNITY_LINE}
- Instagram: {INSTAGRAM_HANDLE}
- Instagram URL: {INSTAGRAM_URL}
- WhatsApp community: {WHATSAPP_COMMUNITY}
- Location: Cape Town, South Africa

## Core Positioning

The Vault Room is a community-driven anime and trading-card collector hideout where every grail has a story. It brings collectors together to buy, sell, trade, consign, battle, rip packs, run trade days, and celebrate rare finds.

## Brand Lines

- Unlock your grail.
- Every card has a story.
- Buy. Sell. Trade. Consign.
- More than cards. We're a community.

## Visual Direction

- Crest: circular vault-keyhole mark with gold ring, blue ocean waves, cream base, and collectible-world symbols.
- Palette: vault navy `#0D4EA2`, blue `#2176D2`, gold `#D4AF37`, coral `#FF6B5B`, sky `#7EC6F0`, cream `#FFF7E6`.
- Tone: premium, warm, collector-first, community-led, adventurous, trustworthy.
- Public safety: use original Vault Room, card, vault, Cape Town, ocean, map, and collector motifs on public pages. Keep fan assets and direct franchise references for gated/private contexts only.

## Brand Assets

- Poster example: `media/branding/vault-room-poster-example.jpg`
- Crest: `media/branding/vault-room-crest.jpg`
- Primary logo: `media/branding/vault-room-primary-logo.jpg`
- Brand board: `media/branding/vault-room-brand-board.jpg`

## Website Revamp Note

Do not update the current app from this note alone. The user explicitly said to record the context first and revamp the website later.
""",
    )
    write_markdown(
        "inventory/inventory_policy.md",
        f"""
# Inventory Policy

Updated: {TODAY}

## Bootstrap Import Status

- status: imported
- source: `/Users/mx/Downloads/export.csv`
- raw copy: `inventory/raw/export.csv`
- generated files: `inventory/master_inventory.csv`, `inventory/master_inventory.xlsx`

## Baseline

```json
{summary_json}
```

## Standing Assumptions

- All CSV rows are physically in hand.
- No CSV rows are inbound unless separately marked.
- No CSV rows are consignment unless separately added.
- No CSV rows are watch-only unless separately marked.
- Zero cost means cost-basis unknown or box pull, not free.
- Existing CSV market price is treated as ZAR by default.
- Fresh researched comps must be stored in USD and ZAR.
- Sealed products are complete sealed products unless stated otherwise.
- Premium Card Collection -Best Selection Vol. 4-, Premium Card Collection -Live Action Edition-, Learn Together Deck Set, boosters, and similar sealed products stay sealed unless break-value analysis is explicitly requested.
- Consignment cards from Drive are separate from CSV inventory.
- Consignment baseline is 80% of researched market value unless changed.
""",
    )
    write_markdown(
        "inventory/inventory_schema.md",
        """
# Inventory Schema

`inventory/master_inventory.csv` is the canonical business spreadsheet generated from Collectr exports and later research.

Important status fields:

- `ownership_status`: `in_hand`, `inbound`, `consignment`, `sold`, `watch`.
- `inventory_status`: `available_for_pricing`, `listed`, `reserved`, `sold`, `hold`, `grading_candidate`, `submitted_for_grading`.
- `cost_status`: `recorded`, `unknown_or_box_pull`.
- `confidence_rating`: `pending_research`, `low`, `medium`, `high`.

Pricing fields stay blank until researched through `pricing/pricing_protocol.md`.
""",
    )
    write_markdown(
        "memory/memory_index.md",
        """
# Memory Index

- `brand_context.md`: The Vault Room public identity, links, palette, brand lines, and app-revamp constraint.
- `business_context.md`: The Vault Room / OP CPT mission, channels, positioning, and rules.
- `inventory_context.md`: current Collectr export baseline and inventory assumptions.
- `pricing_context.md`: Vault Room pricing protocol and example card.
- `market_context.md`: One Piece market research summary and source hierarchy.
- `website_context.md`: current Next.js OP CPT app and catalogue roadmap.
- `agent_council_context.md`: internal council roles for pricing and operations.
""",
    )
    write_markdown(
        "memory/business_context.md",
        f"""
# Business Context

The Vault Room is the public Cape Town collector-community brand. OP CPT remains the internal workspace name and existing app/repo continuity label.

Channels: Cape Town local deals, WhatsApp, Instagram, trade shows, Vault Room trade days, Vault Room tournaments, online sales, future website catalogue, and possible international channels such as eBay, CardTrader, or Cardmarket.

Live site: {LIVE_SITE}
Instagram: {INSTAGRAM_HANDLE}
WhatsApp community: {WHATSAPP_COMMUNITY}

Brand posture: premium, trusted, collector-first, research-backed, local, and community-driven.
""",
    )
    write_markdown(
        "memory/brand_context.md",
        f"""
# Brand Context

Updated: {TODAY}

Public brand: {PUBLIC_BRAND}
Internal workspace continuity name: {INTERNAL_WORKSPACE}

Tagline: {BRAND_TAGLINE}
Community line: {BRAND_COMMUNITY_LINE}

Instagram: {INSTAGRAM_HANDLE}
Instagram URL: {INSTAGRAM_URL}
WhatsApp community: {WHATSAPP_COMMUNITY}

Brand lines:

- Unlock your grail.
- Every card has a story.
- Buy. Sell. Trade. Consign.
- More than cards. We're a community.

Palette:

- vault navy: `#0D4EA2`
- vault blue: `#2176D2`
- vault gold: `#D4AF37`
- vault coral: `#FF6B5B`
- vault sky: `#7EC6F0`
- vault cream: `#FFF7E6`

Local brand assets:

- `media/branding/vault-room-poster-example.jpg`
- `media/branding/vault-room-crest.jpg`
- `media/branding/vault-room-primary-logo.jpg`
- `media/branding/vault-room-brand-board.jpg`

Constraint: user explicitly said not to change the current web/app yet. Record this identity now and use it for the later full website revamp.
""",
    )
    write_markdown(
        "memory/inventory_context.md",
        f"""
# Inventory Context

Updated: {TODAY}

The current Collectr export was imported from `/Users/mx/Downloads/export.csv`.

Summary:

```json
{summary_json}
```

Rows are treated as physically in hand. Consignment Drive items are separate until individually imported.
""",
    )
    write_markdown(
        "memory/pricing_context.md",
        """
# Pricing Context

Use `pricing/pricing_protocol.md` as the standing method.

Hard rules:

- Never price from Collectr alone.
- Card Ladder is required for serious One Piece pricing when accessible.
- Identify exact variant before quoting.
- Quote serious prices in USD and ZAR.
- Build quick-sale, fair-market, patient-local, event/show, online-gross, stretch, and walk-away prices.

Example to remember:

- Monkey.D.Luffy
- P-033
- P rarity / promo
- English
- Event Pack Vol. 2 style promo
- Raw, ungraded
- Owner-described as perfect mint / strong grading candidate

When asked to price it, research current comps again rather than using old cached numbers.

## Prior Price Bible Snapshot

`market_research/price_bible_snapshot_2026-06-12.md` contains a prior show-pricing deliverable. Treat it as historical operating context, not current market truth. Refresh top-card comps before quoting.
""",
    )
    write_markdown(
        "memory/market_context.md",
        """
# Market Context

One Piece TCG is a hybrid player-and-collector market. OP CPT should separate:

- Playable singles and meta-driven demand.
- Collector demand for Luffy, Zoro, Shanks, Ace, Law, Boa, Nami, Sabo, Roger, Mihawk, Yamato, Robin, Chopper, and iconic villains.
- Promo/prize/tournament scarcity.
- Manga rare and SP/chase-card liquidity.
- English versus Japanese pricing and availability.
- Cape Town local scarcity and inspection premium.

Official OP-16 The Time of Battle released June 12, 2026, making current set-cycle demand and event cards especially time-sensitive.
""",
    )
    write_markdown(
        "memory/website_context.md",
        f"""
# Website Context

Live: {LIVE_SITE}

Local app: `apps/op-cpt`

Stack: Next.js App Router, React, React Three Fiber, Supabase-ready API routes, Vercel primary deploy, Netlify fallback.

Current public shell already exposes navigation for Harbor, Collection, Scanner, Trades, Market, Events, Rankings, Group Buys, News, and Admin.

Next commercial step: enquiry-only catalogue with public-safe scans, ZAR pricing, condition, language, status, WhatsApp/Instagram contact, and private admin fields.
""",
    )
    write_markdown(
        "memory/agent_council_context.md",
        """
# Agent Council Context

Use this internal council for serious pricing and buying decisions:

1. Captain / Business Operator
2. Card Identifier
3. Market Comp Analyst
4. Fraud and Bad Data Filter
5. Grading Analyst
6. Scarcity and Market-Making Analyst
7. Cape Town Local Market Analyst
8. Portfolio CFO
9. Acquisition and Bidding Analyst
10. Website Architect
11. Content and Brand Strategist
12. Risk / Legal / Reputation Analyst

Return concise conclusions and evidence. Do not expose private hidden reasoning.
""",
    )


def write_pricing_docs() -> None:
    protocol = (ROOT.parent / ".codex").exists()
    attachment_path = Path("/Users/mx/.codex/attachments/8289567b-d5a4-40db-b93b-2e0ffcc73181/pasted-text.txt")
    attachment_text = attachment_path.read_text(encoding="utf-8") if attachment_path.exists() else ""
    write_markdown(
        "pricing/pricing_protocol.md",
        attachment_text
        or """
# OP CPT / The Vault Room Pricing Protocol v2

Card Ladder is required for serious One Piece pricing. Never price from Collectr alone.
""",
    )
    write_markdown(
        "pricing/card_pricing_template.md",
        """
# Pricing Recommendation — [Card Name / Number]

## Exact Identification
- Name:
- Number:
- Set/source:
- Language:
- Variant:
- Raw/graded:
- Cert if slabbed:
- Identity confidence:

## Market Evidence
| Source | Signal | Notes |
|---|---:|---|
| TCGplayer / Limitless | | |
| CardTrader / Cardmarket | | |
| PriceCharting | | |
| Card Ladder | | |
| Market Movers / Alt | | |
| eBay sold | | |
| eBay active | | |
| PSA APR / pop | | |
| BGS / CGC / GemRate | | |
| MySlabs / Fanatics / Goldin / Heritage | | |
| Social/local | | |

## Clean Comp Interpretation
- Bad data filtered:
- Realistic market:
- Availability:
- Scarcity:
- Trend:
- Buyer psychology:

## Grading Upside
- PSA 10 potential:
- BGS 10 potential:
- CGC 10 / Pristine potential:
- Grade / sell / hold verdict:

## The Vault Room Price Ladder
| Price Type | USD | ZAR | Purpose |
|---|---:|---:|---|
| Quick-sale floor | | | |
| Fair market | | | |
| Patient local ask | | | |
| Event/show ask | | | |
| Online gross ask | | | |
| Stretch ask | | | |
| Walk-away minimum | | | |

## Final Recommendation
- List at:
- Accept serious offers from:
- Do not go below:
- Expected net by channel:
- Verdict:
- Confidence:
""",
    )
    write_markdown(
        "pricing/acquisition_bid_framework.md",
        """
# Acquisition Bid Framework

Calculate:

- exact identity confidence
- true sellable market value
- patient ask
- expected net
- fees and shipping
- grading upside
- liquidity and time-to-sale
- downside risk
- capital lockup
- ideal bid
- maximum bid
- walk-away bid

Formula:

```text
max_bid = expected_net_sell_price - desired_profit - risk_buffer - fees - shipping - grading_costs_if_any
```

For consignment:

```text
consignment_base = researched_market_value * 0.80
```
""",
    )
    write_markdown(
        "pricing/consignment_pricing_framework.md",
        """
# Consignment Pricing Framework

Default baseline: owner receives 80% of researched sellable market value unless another deal is agreed.

Accept consignment only when:

- identity is clear
- condition can be verified
- price spread supports OP CPT effort
- card adds content, customer acquisition, or catalogue strength
- owner accepts realistic timelines and walk-away rules

Do not use Collectr as the market value. Research the card first.
""",
    )
    write_markdown(
        "pricing/grading_ev_framework.md",
        """
# Grading EV Framework

Grade when expected net graded value beats raw patient-sale value after:

- grading fee
- shipping and insurance
- turnaround time
- grade risk
- cashflow lockup
- platform fees

Score raw candidates on centering, corners, edges, surface, print quality, foil quality, and back condition.

Never advertise a raw card as a guaranteed 10.
""",
    )
    write_markdown(
        "pricing/negotiation_framework.md",
        """
# Negotiation Framework

Pricing ladder:

- stretch ask: ambitious market-maker price for rare/low-supply/high-upside cards
- patient ask: main listing price, high but defensible
- event ask: local show/trade-day sticker with negotiation room
- fair market: clean-comp-supported sellable price
- quick-sale floor: fast sale/dealer price
- walk-away minimum: lowest acceptable price unless strategy changes

Use scarcity, local availability, condition, grading upside, and trust as anchors.
""",
    )
    write_markdown(
        "pricing/channel_fee_model.md",
        """
# Channel Fee Model

Track gross, fees, shipping, payment costs, returns/dispute risk, and expected net by channel.

Channels:

- Local cash/EFT: lowest fee, highest trust, limited audience.
- WhatsApp/Instagram: patient ask, relationship-driven, manual payment.
- Trade day/event: sticker premium, bundle negotiation.
- eBay: global demand but high fees, shipping, disputes, and return risk.
- CardTrader/Cardmarket: international benchmark; access and logistics vary from South Africa.
- TCGplayer: benchmark only unless selling access is available.
""",
    )


def write_market_docs() -> None:
    write_markdown(
        "market_research/one_piece_market_overview.md",
        """
# One Piece TCG Market Overview

Generated: 2026-06-16

## Current Read

- Bandai identifies One Piece Card Game as a major driver of its card-business growth since the July 2022 launch.
- OP-16 The Time of Battle is the current official English release cycle, with official release date June 12, 2026 and 126+1 card types.
- Official restriction updates in April 2026 mean meta demand must be checked before pricing playable singles.
- PSA and GemRate data show grading demand remains high across TCGs, so grading turnaround, fees, and population growth matter for slab strategy.
- Collector pricing is concentrated around manga rares, SPs, treasure/serial cards, event/prize promos, Luffy/Zoro/Shanks/Ace/Sabo/Roger/Mihawk/Nami/Boa/Law/Yamato/Robin demand, and low-pop high-grade slabs.

## OP CPT Implication

Use patient asks for rare, clean, locally scarce cards. Use quick-sale pricing only when capital velocity is the explicit goal.

See `market_research/price_bible_snapshot_2026-06-12.md` for a prior show-pricing snapshot and buyout framing. Refresh its comps before using.
""",
    )
    write_markdown(
        "market_research/set_by_set_guide.md",
        """
# Set-by-Set Guide

Use official product pages and card list as the source of truth before pricing.

## Working Buckets

- OP01-OP03: early-set collector gravity; prioritize iconic leaders, manga rares, and early alt arts.
- OP04-OP08: maturing era; evaluate by character, meta relevance, and special printings.
- OP09-OP13: strong modern chase lane; validate overprinting and grading populations.
- OP14-OP16: current-cycle attention; watch release hype, early price spikes, and errata/restriction effects.
- EB/ST/Premium products: separate sealed-product value from individual-card break value.

Each set-specific price decision should record release date, chase cards, manga/SP/treasure cards, meta leaders, reprint risk, active supply, and grading population.
""",
    )
    write_markdown(
        "market_research/promo_and_prize_card_guide.md",
        """
# Promo and Prize Card Guide

Highest OP CPT priority categories:

- manga rares and true prize/trophy cards
- championship, regional, Treasure Cup, finalist, winner, and participant cards
- event pack promos
- super pre-release and pre-release cards
- serial-numbered cards
- anniversary and premium collection cards
- scarce local Cape Town slabs

Pricing rule: exact stamp/source matters. Participant, finalist, winner, and normal promo versions must never be mixed.
""",
    )
    write_markdown(
        "market_research/sealed_product_strategy.md",
        """
# Sealed Product Strategy

Sealed products stay sealed unless explicitly running a break-value analysis.

Evaluate sealed on:

- MSRP versus current ask
- local scarcity
- shipping/import friction
- reprint risk
- chase-card strength
- event/rip-night utility
- long-term display/collector appeal

Complete sealed products in the current inventory include Premium Card Collection items and Learn Together Deck Set.
""",
    )
    write_markdown(
        "market_research/platform_pricing_guide.md",
        """
# Platform Pricing Guide

## TCGplayer
US retail benchmark. Good for market price, low listing, and liquidity. Verify exact variant.

## CardTrader
Global supply and CardTrader Zero/Ready context. Adjust for fees, shipping, and South African access.

## Cardmarket
EU floor and trend source. EU floors are useful but not equal to Cape Town local sell prices.

## PriceCharting
Historical raw/graded sale snapshots. Always verify SKU and individual sales.

## Card Ladder
Required for serious One Piece pricing, especially high-end slabs, manga rares, prize cards, promos, low-pop cards, and graded market movement.

## eBay
Strong active/sold comp source. Filter fake listings, title stuffing, stock photos, weak sellers, and accepted-offer opacity.

## PSA/BGS/CGC/GemRate
Use for cert verification, population, grade distribution, and grading-volume context.

## Collectr
Portfolio baseline only. Never final sell price.
""",
    )
    write_markdown(
        "market_research/grading_and_population_guide.md",
        """
# Grading and Population Guide

Check PSA, BGS, CGC, and GemRate before pricing serious raw or slabbed cards.

Separate:

- PSA 10
- BGS 9.5
- BGS 10
- BGS Black Label
- CGC Gem Mint 10
- CGC Pristine 10

Grading makes sense when the card is clean, liquidity improves meaningfully, and expected net after grading beats a patient raw sale.
""",
    )
    write_markdown(
        "market_research/meta_and_banlist_watch.md",
        """
# Meta and Banlist Watch

Current official sources to recheck:

- Official restrictions: https://en.onepiece-cardgame.com/rules/restriction/
- Official news: https://en.onepiece-cardgame.com/news/
- Limitless: https://onepiece.limitlesstcg.com/
- OnePiece.gg: https://onepiece.gg/
- Nakama Decks: https://nakamadecks.com/
- Egman Events: https://egmanevents.com/

Playable card pricing should be checked against current English and Japanese/Eastern meta, upcoming set support, and restriction risk.
""",
    )
    write_markdown(
        "market_research/fake_and_scam_filtering.md",
        """
# Fake and Scam Filtering

Red flags:

- wrong language hidden in title
- stock photos for expensive raw cards
- base card titled as alt art/SP/manga
- participant card titled as winner
- fake slab labels or mismatched certs
- low-feedback seller with too-good pricing
- relisted auctions that did not really sell
- old comps before reprint/meta/banlist shifts
- Collectr or PriceCharting SKU mismatch

Require front/back scans for high-end raw cards and cert verification for slabs.
""",
    )
    write_markdown(
        "market_research/cape_town_market_notes.md",
        """
# Cape Town Market Notes

Local premium factors:

- buyer can inspect the exact card
- no international shipping
- no customs/VAT surprise
- no fake-listing risk
- no long wait
- local trust and repeat relationship
- trade-day excitement and bundle opportunities

Use global comps as evidence, then adjust for local scarcity and trust.
""",
    )
    write_markdown(
        "market_research/agent_council_context.md",
        (ROOT / "memory/agent_council_context.md").read_text(encoding="utf-8"),
    )


def write_website_content_docs() -> None:
    write_markdown(
        "website/website_context.md",
        f"""
# Website Context

Live site: {LIVE_SITE}

Framework: Next.js App Router under `apps/op-cpt`.

Current routes include public shell, playable harbor, collection, scanner, trades, market, events, rankings, group buys, news, admin, and API routes for cards, prices, trades, events, rankings, and imports.

Current priority: enquiry-only public catalogue. No checkout, escrow, shipping automation, or payments until explicitly requested.
""",
    )
    write_markdown(
        "website/public_catalogue_roadmap.md",
        """
# Public Catalogue Roadmap

MVP:

- public card listing
- own scan/photo
- price in ZAR with optional USD reference
- condition, language, variant, raw/slab
- status: available, reserved, sold, hold
- enquiry button to WhatsApp/Instagram
- source-backed notes where useful

Admin needs:

- import from master inventory
- hide private fields
- mark reserved/sold
- upload scans
- create pricing decision records
""",
    )
    write_markdown(
        "website/data_integration_plan.md",
        """
# Data Integration Plan

Short term:

- Generate catalogue seed data from `inventory/master_inventory.csv`.
- Keep researched prices in `inventory/pricing_research_log.csv`.
- Store public scans separately from private intake images.

Later:

- Supabase tables for cards, inventory, scans, price comps, pricing decisions, sales, purchases, consignments, grading submissions, events, and content posts.
""",
    )
    write_markdown(
        "website/supabase_schema_plan.md",
        """
# Supabase Schema Plan

Future tables:

- cards
- inventory_items
- inventory_scans
- price_comps
- pricing_decisions
- sales
- purchases
- consignments
- consignment_owners
- grading_submissions
- events
- trade_day_listings
- content_posts
- watchlist_items

Public fields: card name, card number, set, rarity, language, condition, price, status, public notes, front image.

Private fields: cost basis, seller/source, consignment owner, minimum price, cert/private notes, profit margin, negotiation history.
""",
    )
    write_markdown(
        "website/vercel_deployment_notes.md",
        """
# Vercel Deployment Notes

Primary deployment: Vercel.

Live URL: https://op-cpt.vercel.app

Before deploying:

```bash
cd apps/op-cpt
npm run lint
npm run typecheck
npm run build
```

Do not expose service-role keys, private scans, owner data, customer details, payment details, or private cost basis in browser bundles.
""",
    )
    write_markdown(
        "content/instagram_strategy.md",
        f"""
# Instagram Strategy

Position The Vault Room as Cape Town's trusted collector community: grail access, transparent pricing, community trading, game nights, grading candidates, and local trust.

Primary handle: {INSTAGRAM_HANDLE}
Community join link: {WHATSAPP_COMMUNITY}

Content pillars:

- new arrivals
- rare card spotlight
- market watch
- grading candidate spotlight
- behind-the-scenes scanning
- trade-day announcements
- tournament announcements
- fake listing warnings
- consignment highlights
""",
    )
    write_markdown(
        "content/content_calendar.md",
        """
# Content Calendar

Weekly rhythm:

- Monday: market watch
- Tuesday: card spotlight
- Wednesday: grading candidate / condition education
- Thursday: trade-day or group-buy update
- Friday: new stock / weekend availability
- Weekend: event, pack rip, or community recap
""",
    )
    write_markdown(
        "content/post_templates.md",
        """
# Post Templates

## New Stock Drop

New arrival at The Vault Room: [card].

- Set/card number:
- Condition/grade:
- Why collectors care:
- Price:
- DM to reserve or inspect in Cape Town.

## Market Watch

Market watch: [card/category].

What changed:
Why it matters:
Vault Room view:
""",
    )
    write_markdown(
        "content/card_spotlight_template.md",
        """
# Card Spotlight Template

Card:
Number:
Set/source:
Language:
Variant:
Condition/grade:
Why it matters:
Market context:
Vault Room price:
Inspection/pickup:
""",
    )
    write_markdown(
        "content/trade_day_content.md",
        """
# Trade Day Content

Key messages:

- Bring binders.
- Bring deck boxes.
- In-person inspection encouraged.
- No pressure trades.
- Rare cards can be reviewed through the Vault Room pricing protocol.
- The Vault Room can help with pricing, grading decisions, and consignment intake.
""",
    )


def write_agent_file() -> None:
    existing = (ROOT / "AGENTS.md").read_text(encoding="utf-8") if (ROOT / "AGENTS.md").exists() else ""
    addition = """

## OP CPT Business Desk

For OP CPT / The Vault Room pricing, inventory, consignment, content, and website work:

- Use `pricing/pricing_protocol.md` before quoting prices.
- Never price from Collectr alone.
- Card Ladder is required for serious One Piece pricing when accessible.
- Save pricing evidence to `inventory/pricing_research_log.csv`.
- Use `inventory/master_inventory.csv` and `inventory/master_inventory.xlsx` as the operating inventory.
- Preserve raw exports in `inventory/raw/`.
- Keep secrets and payment/customer data out of commits.
- Public catalogue work lives under `apps/op-cpt` and `website/`.
"""
    if "## OP CPT Business Desk" not in existing:
        write_markdown("AGENTS.md", existing.rstrip() + addition)


def main() -> None:
    ensure_dirs()
    rows = read_export()
    master, summary = build_master(rows)
    write_csv(ROOT / "inventory/master_inventory.csv", master, MASTER_COLUMNS)
    write_xlsx(master, summary)
    (ROOT / "inventory/baseline_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    write_logs()
    write_docs(summary)
    write_pricing_docs()
    write_market_docs()
    write_website_content_docs()
    write_agent_file()
    if PRICE_BIBLE_REPORT.exists():
        normalized_report = "\n".join(
            line.rstrip() for line in PRICE_BIBLE_REPORT.read_text(encoding="utf-8").splitlines()
        )
        (ROOT / "market_research/price_bible_snapshot_2026-06-12.md").write_text(
            normalized_report + "\n",
            encoding="utf-8",
        )
    write_markdown(
        "scripts/README.md",
        """
# Scripts

- `bootstrap_op_cpt_business.py`: refreshes OP CPT business docs and inventory from `/Users/mx/Downloads/export.csv`.
""",
    )


if __name__ == "__main__":
    main()
